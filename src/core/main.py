import sys
from PySide6.QtCore import Qt
from PySide6.QtGui import QFont, QIcon
from os import environ, path
from PySide6.QtGui import QShortcut, QKeySequence, QTextCursor
from PyPDF2 import PdfReader
from docx import Document
import os

from PySide6.QtWidgets import (
    QApplication,
    QMainWindow,
    QVBoxLayout,
    QHBoxLayout,
    QWidget,
    QPushButton,
    QPlainTextEdit,
    QFileDialog,
    QLabel,
    QColorDialog,
    QFontDialog,
    QSlider,
    QInputDialog,
    QMessageBox
    
)

env = environ.get("ENV", "production")

class CodeEditor(QMainWindow):
    def __init__(self):
        super().__init__()
        
        if env == "development":
            icon_path = path.join(
                path.dirname(__file__),
                "..",
                "assets",
                "images",
                "icons",
                "orange.ico",
            )
        else:
            icon_path = path.join(
                path.dirname(__file__), "assets", "images", "icons", "orange.ico"
            )

        self.setWindowIcon(QIcon(icon_path))
        self.setWindowTitle("Orange Editor")
        self.setGeometry(100, 100, 800, 600)

        self.font = QFont("Consolas", 10)
        self.setFont(self.font)

        self.init_ui()

    def init_ui(self):
        """Inicializa os componentes da interface gráfica."""
        self.setStyleSheet("background-color: #1E1E1E; color: #D4D4D4;")
        self.create_shortcuts()
        
        self.editor = QPlainTextEdit(self)
        self.editor.setFont(self.font)
        self.editor.setStyleSheet("background-color: #1E1E1E; color: #D4D4D4;")
        self.editor.setTabStopDistance(4 * self.font.pointSizeF())
        self.editor.verticalScrollBar().valueChanged.connect(self.sync_line_number_scroll)
        self.editor.textChanged.connect(self.update_line_numbers)

        self.line_number_area = QWidget(self)
        self.line_number_area.setStyleSheet("background-color: #2E2E2E; color: #888888;")
        self.line_number_layout = QVBoxLayout(self.line_number_area)
        self.line_number_layout.setContentsMargins(0, 0, 0, 0)
        self.line_number_layout.setAlignment(Qt.AlignTop)
        self.line_number_labels = []

        self.open_button = QPushButton("Abrir", self)
        self.open_button.setStyleSheet("background-color: #333333; color: #FFFFFF;")
        self.open_button.clicked.connect(self.open_file)

        self.new_button = QPushButton("Novo", self)
        self.new_button.setStyleSheet("background-color: #333333; color: #FFFFFF;")
        self.new_button.clicked.connect(self.new_file)

        self.save_button = QPushButton("Salvar", self)
        self.save_button.setStyleSheet("background-color: #333333; color: #FFFFFF;")
        self.save_button.clicked.connect(self.save_file)

        self.format_button = QPushButton("Cor", self)
        self.format_button.setStyleSheet("background-color: #333333; color: #FFFFFF;")
        self.format_button.clicked.connect(self.open_format_dialog)

        self.transparent_button = QPushButton("Alternar Transparência", self)
        self.transparent_button.setStyleSheet("background-color: #333333; color: #FFFFFF;")
        self.transparent_button.clicked.connect(self.toggle_transparency)

        self.transparency_slider = QSlider(Qt.Horizontal, self)
        self.transparency_slider.setMinimum(0)
        self.transparency_slider.setMaximum(100)
        self.transparency_slider.setValue(100) 
        self.transparency_slider.setTickInterval(10)
        self.transparency_slider.setTickPosition(QSlider.TicksBelow)
        self.transparency_slider.valueChanged.connect(self.update_transparency)

        buttons_layout = QHBoxLayout()
        buttons_layout.addWidget(self.open_button)
        buttons_layout.addWidget(self.save_button)
        buttons_layout.addWidget(self.new_button)
        buttons_layout.addWidget(self.format_button)
        buttons_layout.addWidget(self.transparent_button) 

        editor_layout = QHBoxLayout()
        editor_layout.addWidget(self.line_number_area)
        editor_layout.addWidget(self.editor)

        transparency_layout = QVBoxLayout()
        transparency_layout.addWidget(QLabel("Ajustar Transparência:"))
        transparency_layout.addWidget(self.transparency_slider)
        
        main_layout = QVBoxLayout()
        main_layout.addLayout(buttons_layout)
        main_layout.addLayout(editor_layout)
        main_layout.addLayout(transparency_layout)

        central_widget = QWidget(self)
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)

        self.transparent = False
        self.update_line_numbers()

    def update_line_numbers(self):
        """Atualiza a exibição de números de linha."""
        # Limpa os números existentes
        for label in self.line_number_labels:
            self.line_number_layout.removeWidget(label)
            label.deleteLater()
        self.line_number_labels = []


    def sync_line_number_scroll(self, value):
        """Sincroniza a rolagem da área de números de linha com a rolagem do editor."""
        self.line_number_area.move(1, -value) 

    def create_shortcuts(self):
        """Cria atalhos para os botões.""" 
        self.open_shortcut = QShortcut(QKeySequence("Ctrl+O"), self)
        self.open_shortcut.activated.connect(self.open_file)

        self.save_shortcut = QShortcut(QKeySequence("Ctrl+S"), self)
        self.save_shortcut.activated.connect(self.save_file)

        self.new_shortcut = QShortcut(QKeySequence("Ctrl+N"), self)
        self.new_shortcut.activated.connect(self.new_file)

        self.format_shortcut = QShortcut(QKeySequence("Ctrl+F"), self)
        self.format_shortcut.activated.connect(self.open_format_dialog)

        self.transparent_shortcut = QShortcut(QKeySequence("Ctrl+T"), self)
        self.transparent_shortcut.activated.connect(self.toggle_transparency)

        self.goto_shortcut = QShortcut(QKeySequence("Ctrl+G"), self)
        self.goto_shortcut.activated.connect(self.goto_line)

    def goto_line(self):
        """Abre um diálogo para o usuário inserir o número da linha."""
        line, ok = QInputDialog.getInt(self, "Ir para a Linha", "Digite o número da linha:", 1, 1, self.editor.blockCount(), 1)
        if ok:
            self.jump_to_line(line)

    def jump_to_line(self, line_number):
        """Move o cursor para o número da linha especificado e seleciona a linha inteira."""
        block = self.editor.document().findBlockByNumber(line_number - 1)
        cursor = self.editor.textCursor()
        cursor.setPosition(block.position())
        
        cursor.movePosition(QTextCursor.EndOfBlock, QTextCursor.KeepAnchor)
        
        self.editor.setTextCursor(cursor)

    

    def update_line_area(self, rect, dy):
        """Sincroniza os números de linha com o editor.""" 
        if dy:
            self.line_number_area.scroll(0, dy)
        else:
            self.line_number_area.update()

    def new_file(self):
        """Cria um novo arquivo e limpa o editor.""" 
        self.editor.setPlainText("")

    def open_file(self):
        """Abre um arquivo e carrega o conteúdo no editor."""
        file_path, _ = QFileDialog.getOpenFileName(self, "Abrir Arquivo", "", "Todos os Arquivos (*.*)")
        if file_path:
            ext = os.path.splitext(file_path)[-1].lower()
            try:
                if ext in [".txt", ".gitignore", ".log", ".cfg"]:  
                    with open(file_path, "r", encoding="utf-8") as file:
                        content = file.read()
                        self.editor.setPlainText(content)

                elif ext == ".docx":  
                    from docx import Document
                    doc = Document(file_path)
                    content = "\n".join([p.text for p in doc.paragraphs])
                    self.editor.setPlainText(content)

                elif ext == ".xlsx":  
                    from openpyxl import load_workbook
                    workbook = load_workbook(file_path)
                    sheet = workbook.active
                    rows = [[str(cell.value) for cell in row] for row in sheet.iter_rows()]
                    content = "\n".join(["\t".join(row) for row in rows])
                    self.editor.setPlainText(content)

                elif ext == ".pdf":  
                    from PyPDF2 import PdfReader
                    reader = PdfReader(file_path)
                    content = "\n".join([page.extract_text() for page in reader.pages])
                    self.editor.setPlainText(content)

                else:  
                    with open(file_path, "rb") as file:
                        binary_data = file.read()
                    try:
                        decoded_text = binary_data.decode("utf-8")
                        formatted_text = self.format_content(decoded_text, ext)
                        self.editor.setPlainText(formatted_text)
                    except UnicodeDecodeError:
                        hex_view = binary_data.hex()
                        self.editor.setPlainText(f"Arquivo binário detectado. Exibindo em hexadecimal:\n\n{hex_view}")

            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Não foi possível abrir o arquivo:\n{e}")

    def format_content(self, content, ext):
        """Aplica a formatação apropriada ao conteúdo com base na extensão."""
        if ext == ".html":
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(content, "html.parser")
            return soup.prettify()
        
        elif ext == ".json":
            import json
            try:
                parsed = json.loads(content)
                return json.dumps(parsed, indent=4)
            except json.JSONDecodeError:
                return content  # Retorna o texto original se não puder ser formatado

        # Adicionar mais casos de formatação, se necessário
        return content
            

    def save_file(self):
        """Salva o conteúdo do editor em um arquivo.""" 
        file_path, _ = QFileDialog.getSaveFileName(self, "Salvar Arquivo", "", "Todos os Arquivos (*.*)")
        if file_path:
            with open(file_path, "w", encoding="utf-8") as file:
                content = self.editor.toPlainText()
                file.write(content)

    def open_format_dialog(self):
        """Abre um diálogo para personalizar o estilo do editor.""" 
        color = QColorDialog.getColor()
        if color.isValid():
            self.editor.setStyleSheet(
                f"background-color: #1E1E1E; color: {color.name()};"
            )

    def toggle_transparency(self):
        """Alterna a transparência da janela.""" 
        if self.transparent:
            self.setWindowOpacity(1) 
        else:
            self.setWindowOpacity(0.5)  
        
        self.transparent = not self.transparent

    def update_transparency(self):
        """Atualiza a transparência conforme o valor do slider.""" 
        opacity_value = self.transparency_slider.value() / 100  
        self.setWindowOpacity(opacity_value)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = CodeEditor()
    window.show()
    sys.exit(app.exec())
